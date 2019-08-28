from tkinter.filedialog import askopenfilename
from tkinter import *
import tkinter.messagebox
from tkinter import ttk

from openpyxl import load_workbook
import smtplib
import re
import csv

from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 




root=Tk()
root.geometry("640x340")

root.title("Email From Excel")

root.configure(background="light green")

root.iconbitmap("email.ico")

label4=Label(root,text="Attach File -->",bg="light green")#Attach File
label4.place(x=490,y=317)

img=PhotoImage(file="attach.png")#Photo of attach file
img=img.subsample(20,35)



label5=Label(root,text="Subject-->",bg="light green")
label5.place(x=400,y=120)#Subject label

val3=StringVar()#Subject variable
entry3=ttk.Entry(root,textvariable=val3)#Subject Field
entry3.place(x=490,y=120)

ee=""

attacher=""

def attach():
    global attacher
    attacher=askopenfilename()
    tkinter.messagebox.showinfo("Attachment:",("The selected attachment is",attacher))

button3=ttk.Button(root,image=img,command=attach)#Attach file button
button3.place(x=579,y=315)


def opendir():
    filename=askopenfilename()
    file=filename
    if filename.endswith("xlsx"):
     
     x = load_workbook(filename)
     s = x.active
     a =[]
     m_row = s.max_row
     m_col = s.max_column
     for i in range(1,m_row+1):
      for j in range(1,m_col+1):
        cell_obj = s.cell(row = i , column =j )
        v = cell_obj.value
        a.append(str(v))
     p = " ".join(a)

     
     e = re.findall('([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', p)
     global ee
     ee=e
     
     mail=len(ee)

     label3=Label(root,text=file,font="times 12")
     label3.place(x=210,y=85)
     tkinter.messagebox.showinfo("Emails Selected:",("Mails selected",mail))
     

    elif filename.endswith("csv"):
        a=[]
        count=0
        print(filename)
        file=filename
        with open(filename, 'r') as csvFile:
            reader = csv.reader(csvFile)
            for i in reader:
                for j in i:
                 
                 a.append(j)
            ee=a
            
            for i in a:
             
             if i.endswith("@gmail.com"):
                    count+=1
            tkinter.messagebox.showinfo("Emails Selected:",("Mails selected",count))
        
     


            label3=Label(root,text=file,font="times 12")
            label3.place(x=210,y=85)
              



def Login(*args):
    
     host = "smtp.gmail.com"
     port = 465
     username = val1.get() #"pythontutorials46@gmail.com"
     password = val2.get()
     subject=val3.get()
     msgg=text1.get("1.0","end-1c")
     from_mail = username
   
# instance of MIMEMultipart
     msg = MIMEMultipart() 
  
# storing the senders email address   
     msg['From'] = from_mail
  
# storing the receivers email address  
 
  
# storing the subject  
     msg['Subject'] = subject
  
# string to store the body of the mail 
     body = msgg
  
# attach the body with the msg instance 
     msg.attach(MIMEText(body, 'plain')) 
  
# open the file to be sent  
     filename = attacher
     attachment = open(filename, "rb")
  
# instance of MIMEBase and named as p 
     p = MIMEBase('application', 'octet-stream') 
  
# To change the payload into encoded form 
     p.set_payload((attachment).read()) 
  
# encode into base64 
     encoders.encode_base64(p) 
   
     p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
  
# attach the instance 'p' to instance 'msg' 
     msg.attach(p) 
  

  
# Converts the Multipart msg into a string 
     text = msg.as_string() 
  
# sending the mail 


     
     
     
     try:
       
      email_conn = smtplib.SMTP_SSL(host,port)
      email_conn.ehlo()
      
      email_conn.login(username, password)
      for i in ee:
          if i.endswith("@gmail.com"):
           email_conn.sendmail(from_mail, i ,text)
        
      email_conn.quit()

        
      tkinter.messagebox.showinfo("Successful !!","Message Successfully sent")
      
     except:
      tkinter.messagebox.showinfo("Failed !!","Message sending failed")
      print("There is an error in connecting to the server")

val1=StringVar()
val2=StringVar()



label1=ttk.Label(root,text="From_Name")
label1.config(background="light green")

label2=ttk.Label(root,text="Password")
label2.config(background="light green")


entry1=ttk.Entry(root,textvariable=val1)
entry2=ttk.Entry(root,textvariable=val2)
entry2.config(show="*")

label1.place(x=10,y=150)
label2.place(x=25,y=180)


entry1.place(x=100,y=150)
entry2.place(x=100,y=180)

text1=Text(root,height=10,width=40)
text1.place(x=290,y=150)


button1=Button(root,text="Choose Files",
               command=opendir,
               padx=20,
               pady=10,
               fg="black",
               bg="light blue",
               font=("times",16,"bold")
               )
button1.place(x=250,y=10)

button2=ttk.Button(root,text="SEND",

               
               command=Login
               )
button2.place(x=80,y=220)



root.mainloop()
